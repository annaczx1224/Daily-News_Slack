import os
import sys
import sqlite3
import uuid

import openpyxl


# --------------------------------------------------
# PROJECT PATHS
# --------------------------------------------------

PROJECT_ROOT = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        ".."
    )
)

SRC_PATH = os.path.join(
    PROJECT_ROOT,
    "src"
)

sys.path.append(SRC_PATH)


from entity_resolution import normalize_company_name


DATABASE_PATH = os.path.join(
    PROJECT_ROOT,
    "deal_intelligence.db"
)

EXCEL_PATH = os.path.join(
    PROJECT_ROOT,
    "Enterprise Tech.xlsx"
)


# Your PitchBook export has its headers on row 10.
HEADER_ROW = 10

COMPANY_COLUMN = "Companies"


# --------------------------------------------------
# HELPERS
# --------------------------------------------------

def generate_company_id():
    """
    Generate a SenaHill-owned permanent company ID.

    Example:
    cmp_17fd8aa71c78448c...
    """

    return "cmp_" + uuid.uuid4().hex


def get_column_indexes(sheet):
    """
    Read the actual headers from row 10 and
    return a dictionary mapping header -> column number.
    """

    headers = {}

    for cell in sheet[HEADER_ROW]:

        if cell.value is None:
            continue

        header = str(cell.value).strip()

        headers[header] = cell.column

    return headers


# --------------------------------------------------
# LOAD COMPANIES
# --------------------------------------------------

def load_companies():

    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(
            f"Excel file not found: {EXCEL_PATH}"
        )

    if not os.path.exists(DATABASE_PATH):
        raise FileNotFoundError(
            f"Database not found: {DATABASE_PATH}\n"
            f"Run sql/schema.sql first."
        )

    print("Opening Excel file...")

    workbook = openpyxl.load_workbook(
        EXCEL_PATH,
        read_only=True,
        data_only=True
    )

    sheet = workbook.active

    headers = get_column_indexes(sheet)

    print("\nColumns found:")

    for header in headers:
        print(f" - {header}")

    if COMPANY_COLUMN not in headers:
        raise ValueError(
            f"Could not find '{COMPANY_COLUMN}' column."
        )

    company_col = headers[COMPANY_COLUMN]

    conn = sqlite3.connect(DATABASE_PATH)

    cursor = conn.cursor()

    inserted = 0
    skipped = 0
    blank = 0

    # Start immediately below the header row.
    for row_number in range(
        HEADER_ROW + 1,
        sheet.max_row + 1
    ):

        company_name = sheet.cell(
            row=row_number,
            column=company_col
        ).value

        if company_name is None:
            blank += 1
            continue

        company_name = str(company_name).strip()

        if not company_name:
            blank += 1
            continue

        normalized_name = normalize_company_name(
            company_name
        )

        if not normalized_name:
            blank += 1
            continue

        # ------------------------------------------
        # Check whether the exact canonical name
        # already exists.
        #
        # We deliberately DO NOT use normalized_name
        # alone as the duplicate test because two
        # legitimately different companies could
        # normalize to the same generic name.
        # ------------------------------------------

        cursor.execute(
            """
            SELECT company_id
            FROM companies
            WHERE canonical_name = ?
            """,
            (company_name,)
        )

        existing = cursor.fetchone()

        if existing:
            skipped += 1
            continue

        company_id = generate_company_id()

        cursor.execute(
            """
            INSERT INTO companies (
                company_id,
                canonical_name,
                normalized_name,
                universe_flag,
                created_at,
                updated_at
            )
            VALUES (
                ?,
                ?,
                ?,
                1,
                CURRENT_TIMESTAMP,
                CURRENT_TIMESTAMP
            )
            """,
            (
                company_id,
                company_name,
                normalized_name
            )
        )

        inserted += 1

    conn.commit()

    # ------------------------------------------
    # Show final company count
    # ------------------------------------------

    cursor.execute(
        """
        SELECT COUNT(*)
        FROM companies
        """
    )

    total_companies = cursor.fetchone()[0]

    conn.close()
    workbook.close()

    print("\n--------------------------------")
    print("Company universe load complete")
    print("--------------------------------")

    print(f"Inserted: {inserted}")
    print(f"Already existed: {skipped}")
    print(f"Blank / invalid: {blank}")
    print(f"Total companies in database: {total_companies}")


if __name__ == "__main__":
    load_companies()
